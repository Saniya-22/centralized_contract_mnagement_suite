#!/usr/bin/env python3
"""Run regression checks from Response Comparison.xlsx.

Uses the orchestrator in-process by default (no server, no login). Optionally use
--api to hit a running server with login.

Reads the spreadsheet, runs selected queries, and writes:
  1) detailed CSV of all run rows
  2) markdown summary with pass/fail heuristics

Usage (no login — direct orchestrator):
  python3 scripts/regression_from_excel.py \\
    --xlsx "/Users/me/Downloads/Response Comparison.xlsx" \\
    --priority \\
    --out-prefix docs/regression_from_excel

  python3 scripts/regression_from_excel.py \\
    --xlsx "/Users/me/Downloads/Response Comparison.xlsx" \\
    --rows 3,14,16,19,20,2,21,72,5,6,9 \\
    --out-prefix docs/regression_from_excel

With API (requires login):
  python3 scripts/regression_from_excel.py --xlsx "..." --priority --api \\
    --api-url http://localhost:8000 --email you@x.com --password '***' \\
    --out-prefix docs/regression_from_excel
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List

from run_test_queries import _get_token, _path_summary, _post_query

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))


@dataclass
class ExcelRow:
    row_num: int
    query: str
    previous_response: str
    current_response: str
    user_feedback: str
    category: str


PRIORITY_ROWS = [3, 14, 16, 19, 20, 2, 21, 72, 5, 6, 9]


def _categorize(query: str) -> str:
    q = (query or "").lower()
    if any(t in q for t in ("also include", "add clause", "include 52.", "include far")):
        return "letter_amendment"
    if any(t in q for t in ("write ", "draft ", "serial letter", "rea", "rfi")):
        return "letter_request"
    if any(t in q for t in ("commissioning", "punchlist", "closeout", "substantial completion")):
        return "lifecycle"
    if any(t in q for t in ("off site", "off-site", "offsite", "stored materials")):
        return "offsite_storage"
    if "document generator" in q:
        return "product_ui"
    if any(t in q for t in ("gao", "shipyard", "puget sound")):
        return "fact_heavy"
    if re.search(r"\bwhat\s+(is|are)\b", q):
        return "definition"
    return "other"


def _parse_rows_arg(rows: str) -> List[int]:
    out: List[int] = []
    for part in rows.split(","):
        part = part.strip()
        if not part:
            continue
        out.append(int(part))
    return sorted(set(out))


def _load_rows(xlsx_path: str, sheet: str) -> List[ExcelRow]:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("Missing dependency: openpyxl. Install with `pip install openpyxl`.") from e

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet]
    rows: List[ExcelRow] = []
    for r in range(2, ws.max_row + 1):
        q = str(ws.cell(r, 1).value or "").strip()
        if not q:
            continue
        prev = str(ws.cell(r, 2).value or "").strip()
        cur = str(ws.cell(r, 3).value or "").strip()
        fb = str(ws.cell(r, 4).value or "").strip()
        rows.append(
            ExcelRow(
                row_num=r,
                query=q,
                previous_response=prev,
                current_response=cur,
                user_feedback=fb,
                category=_categorize(q),
            )
        )
    return rows


async def _run_direct(query: str) -> Dict[str, Any]:
    from src.agents.orchestrator import GovGigOrchestrator

    orchestrator = GovGigOrchestrator()
    return await orchestrator.run_async(query)


def _run_queries_direct(rows: Iterable[ExcelRow]) -> List[Dict[str, Any]]:
    import asyncio

    out: List[Dict[str, Any]] = []

    async def _run_all():
        for row in rows:
            try:
                result = await _run_direct(row.query)
                out.append({"ok": True, "row": row, "result": result, "error": ""})
            except Exception as e:  # pragma: no cover
                out.append({"ok": False, "row": row, "result": {}, "error": f"{type(e).__name__}: {e}"})

    asyncio.run(_run_all())
    return out


def _run_queries_api(rows: Iterable[ExcelRow], api_url: str, email: str, password: str) -> List[Dict[str, Any]]:
    token = _get_token(api_url, email, password)
    out: List[Dict[str, Any]] = []
    for row in rows:
        try:
            result = _post_query(api_url, token, row.query)
            out.append({"ok": True, "row": row, "result": result, "error": ""})
        except Exception as e:
            out.append({"ok": False, "row": row, "result": {}, "error": f"{type(e).__name__}: {e}"})
    return out


def _eval_case(category: str, result: Dict[str, Any], response: str) -> tuple[str, str]:
    path = _path_summary(result.get("agent_path", []))
    mode = str(result.get("mode") or "").lower()
    response_l = response.lower()
    docs = len(result.get("documents", []))

    if category == "definition":
        if path == "clarifier":
            return "FAIL", "Definition routed to clarifier."
        return "PASS", "Definition answered without clarifier."

    if category == "offsite_storage":
        if path == "clarifier" or mode == "clarify":
            return "FAIL", "Off-site storage should be answered in-scope."
        if docs <= 0:
            return "WARN", "No documents retrieved for off-site storage."
        return "PASS", "In-scope routing with evidence."

    if category == "letter_amendment":
        if "error" in response_l:
            return "FAIL", "Execution error."
        if path == "clarifier":
            return "FAIL", "Amendment should not go to clarifier."
        if not any(t in response_l for t in ("subject:", "dear", "sincerely", "request for equitable adjustment", "serial letter")):
            return "WARN", "Amendment may not have produced letter-form output."
        return "PASS", "Amendment likely handled as letter flow."

    if category == "lifecycle":
        if path == "clarifier":
            return "FAIL", "Lifecycle query routed to clarifier."
        if "contracting officer" in response_l and "commissioning" not in response_l and "punchlist" not in response_l:
            return "WARN", "Lifecycle response may be overly contract/CO-oriented."
        return "PASS", "Lifecycle routing looks acceptable."

    if category == "product_ui":
        if path == "clarifier" or mode in ("copilot", "clarify"):
            return "PASS", "Product/UI query appropriately handled as non-regulation."
        return "WARN", "Product/UI query may have been treated as regulation."

    if category == "fact_heavy":
        if len(response.split()) < 80:
            return "WARN", "Fact-heavy answer seems short; may miss specifics."
        return "PASS", "Fact-heavy response has substantial detail."

    return "INFO", "No strict heuristic for this category."


def _to_csv_rows(records: Iterable[Dict[str, Any]]) -> List[Dict[str, str]]:
    csv_rows: List[Dict[str, str]] = []
    for rec in records:
        row: ExcelRow = rec["row"]
        ok = rec["ok"]
        result = rec["result"] or {}
        response = str(result.get("response") or "").strip()
        verdict, reason = _eval_case(row.category, result, response) if ok else ("FAIL", rec["error"])
        csv_rows.append(
            {
                "row_num": str(row.row_num),
                "category": row.category,
                "query": row.query,
                "path": _path_summary(result.get("agent_path", [])) if ok else "error",
                "mode": str(result.get("mode") or "") if ok else "",
                "doc_count": str(len(result.get("documents", []))) if ok else "0",
                "confidence": str(result.get("confidence") or "") if ok else "",
                "verdict": verdict,
                "reason": reason,
                "response": response if ok else rec["error"],
                "user_feedback": row.user_feedback,
            }
        )
    return csv_rows


def _write_csv(path: str, rows: List[Dict[str, str]]) -> None:
    fields = [
        "row_num",
        "category",
        "query",
        "path",
        "mode",
        "doc_count",
        "confidence",
        "verdict",
        "reason",
        "response",
        "user_feedback",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def _write_summary_md(path: str, rows: List[Dict[str, str]]) -> None:
    counts: Dict[str, int] = {}
    for r in rows:
        counts[r["verdict"]] = counts.get(r["verdict"], 0) + 1

    lines = [
        "# Excel Regression Summary",
        "",
        f"- Total rows run: {len(rows)}",
        f"- PASS: {counts.get('PASS', 0)}",
        f"- WARN: {counts.get('WARN', 0)}",
        f"- FAIL: {counts.get('FAIL', 0)}",
        f"- INFO: {counts.get('INFO', 0)}",
        "",
        "## Fails and Warnings",
        "",
    ]
    for r in rows:
        if r["verdict"] in {"FAIL", "WARN"}:
            lines.append(
                f"- R{r['row_num']} [{r['category']}] {r['verdict']}: {r['reason']} | Query: {r['query']}"
            )
    if all(r["verdict"] not in {"FAIL", "WARN"} for r in rows):
        lines.append("- None.")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def _write_xlsx_new_responses(path: str, records: List[Dict[str, Any]]) -> None:
    """Write an Excel file with original columns plus a 'new_responses' column."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment
    except ImportError as e:
        raise RuntimeError("Missing dependency: openpyxl. Install with `pip install openpyxl`.") from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["User Query", "Previous Response", "Current Response", "User Feedback", "new_responses"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    wrap = Alignment(wrap_text=True, vertical="top")
    for r_idx, rec in enumerate(records, start=2):
        row: ExcelRow = rec["row"]
        result = rec.get("result") or {}
        response_text = (result.get("response") or "").strip() if rec.get("ok") else str(rec.get("error", ""))
        ws.cell(row=r_idx, column=1, value=row.query).alignment = wrap
        ws.cell(row=r_idx, column=2, value=row.previous_response).alignment = wrap
        ws.cell(row=r_idx, column=3, value=row.current_response).alignment = wrap
        ws.cell(row=r_idx, column=4, value=row.user_feedback).alignment = wrap
        ws.cell(row=r_idx, column=5, value=response_text).alignment = wrap
    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser(description="Run regression checks from Response Comparison.xlsx")
    ap.add_argument("--xlsx", required=True, help="Path to the Excel file")
    ap.add_argument("--sheet", default="Sheet1", help="Sheet name (default: Sheet1)")
    ap.add_argument("--rows", help="Comma-separated Excel row numbers to run (e.g. 3,5,6)")
    ap.add_argument("--priority", action="store_true", help="Run predefined high-priority rows")
    ap.add_argument("--api", action="store_true", help="Run via API instead of direct orchestrator")
    ap.add_argument("--api-url", default=os.environ.get("API_BASE_URL", "http://localhost:8000"))
    ap.add_argument("--email", default=os.environ.get("API_TEST_EMAIL"))
    ap.add_argument("--password", default=os.environ.get("API_TEST_PASSWORD"))
    ap.add_argument("--out-prefix", default="docs/regression_from_excel", help="Output file prefix")
    ap.add_argument("--out-xlsx", metavar="FILE", help="Write results to an Excel file with column 'new_responses' (e.g. .../new_responses.xlsx)")
    args = ap.parse_args()

    rows = _load_rows(args.xlsx, args.sheet)
    rows_by_num = {r.row_num: r for r in rows}

    target_nums: List[int]
    if args.rows:
        target_nums = _parse_rows_arg(args.rows)
    elif args.priority:
        target_nums = PRIORITY_ROWS
    else:
        target_nums = [r.row_num for r in rows]

    selected = [rows_by_num[n] for n in target_nums if n in rows_by_num]
    if not selected:
        raise RuntimeError("No matching rows selected.")

    print(f"Loaded {len(rows)} spreadsheet rows; running {len(selected)} selected rows.")
    if args.api:
        print("Using API (login required).")
        if not args.email or not args.password:
            raise RuntimeError("--api requires --email and --password (or API_TEST_EMAIL/API_TEST_PASSWORD)")
        records = _run_queries_api(selected, args.api_url, args.email, args.password)
    else:
        print("Using direct orchestrator (no server, no login).")
        records = _run_queries_direct(selected)

    csv_rows = _to_csv_rows(records)
    out_csv = f"{args.out_prefix}.csv"
    out_md = f"{args.out_prefix}.md"
    os.makedirs(os.path.dirname(out_csv) or ".", exist_ok=True)
    _write_csv(out_csv, csv_rows)
    _write_summary_md(out_md, csv_rows)

    print(f"Wrote CSV: {out_csv}")
    print(f"Wrote summary: {out_md}")

    if args.out_xlsx:
        _write_xlsx_new_responses(args.out_xlsx, records)
        print(f"Wrote Excel (new_responses): {args.out_xlsx}")


if __name__ == "__main__":
    main()
